"""
쓰레드 * 1: 글 목록 새로고침, 클릭할 글 수집
쓰레드 * N: 글 클릭, 소스 수집
쓰레드 * 1: 다운로드
"""

import sys
print(sys.version)
print(sys.executable)

import concurrent.futures

import queue
import time
from datetime import datetime, timedelta
import os

import requests
from bs4 import BeautifulSoup
from urllib.parse import urljoin

from openpyxl import load_workbook

import logging
logging.basicConfig(level=logging.INFO)
dirname = os.path.dirname(__file__)

COLLECTOR_THREADS_NUM = 3
POST_WATCH_RANGE = 10
GALL_URL = "https://gall.dcinside.com/board/lists/?id=baseball_new10"
KOREAN_IPS_XLSX = os.path.join(dirname, "KR_NIPLIST_(2021.07.10).xlsx")
WRITER_TYPES = ["해외유동", "국내유동"]  # "국내유동", "해외유동", "고닉"
SAVE_DIR = "E:/dc_scraper_down"
SHUTDOWN_TIME = ""  # [n.year, n.month, n.day, n.hour, n.minute, n.second]
SHUTDOWN_TIME_DELTA = timedelta(seconds=30)

class ShutdownTimer:
    def __init__(self, *args, t_delta=None):
        logging.info("ShutdownTimer 초기화")
        now = datetime.now()
        if t_delta:
            self.SHUTDOWN_TIME = now + t_delta
        else:
            self.SHUTDOWN_TIME = datetime(*args)

    def is_before_shutdown_time(self):
        if datetime.now() < self.SHUTDOWN_TIME:
            logging.info(f"종료까지 {self.SHUTDOWN_TIME - datetime.now()} 남음")
            return True
        else:
            return False


class IP_Distinguisher:
    def __init__(self, IPS_XLSX):
        self.IPS_XLSX = IPS_XLSX
        self.KR_IP_RANGES = []
        self.parse_ip_ranges()

    def parse_ip_ranges(self):
        load_wb = load_workbook(self.IPS_XLSX, data_only=True)
        load_ws = load_wb['KR_NIPLIST_(2021.07.10)']

        for r in range(3, load_ws.max_row + 1):
            start = load_ws[f'B{r}'].value
            end = load_ws[f'C{r}'].value
            start2 = start.split(".")[0:2]
            str_start2 = "" + start2[0] + "." + start2[1]
            end2 = end.split(".")[0:2]
            str_end2 = "" + end2[0] + "." + end2[1]

            self.KR_IP_RANGES.append([str_start2, str_end2])

    def is_korean_ip(self, ip):
        if not ip:
            logging.warning("아이피가 없습니다")
            raise ValueError
        else:

            logging.debug("iskorean 함수시작")
            for kr_ip_range in self.KR_IP_RANGES:
                logging.debug("for문 iskorean")
                if ip in IP_Distinguisher.IPRange(*kr_ip_range):
                    logging.debug("iskorean True")
                    return True
                logging.debug("iskorean for문 끝")
            return False

    class IPRange:
        def __init__(self, start_ip, end_ip):
            logging.debug("iprange 객체 생성")
            self.start_ip = start_ip
            self.end_ip = end_ip

        def __contains__(self, str_ip):
            starts = [int(i) for i in self.start_ip.split(".")]
            start = starts[0] * 256 + starts[1]

            ends = [int(i) for i in self.end_ip.split(".")]
            end = ends[0] * 256 + ends[1]

            ips = [int(i) for i in str_ip.split(".")]
            ip = ips[0] * 256 + ips[1]

            if start <= ip <= end:
                return True
            else:
                return False


class PostDataGetter:
    def __init__(self, post_tr):
        self.post_tr = post_tr

    def get_post_num(self):
        if self.post_tr.has_attr("data-no"):
            return self.post_tr["data-no"]
        else:
            return ""

    def get_post_type_str(self):
        em = self.post_tr.find("em", class_="icon_img")
        return em["class"][1]

    def get_post_writer(self):
        writer_td = self.post_tr.find(class_='gall_writer ub-writer')
        writer_ip = writer_td["data-ip"]
        writer_nick = writer_td["data-nick"]
        return {"ip": writer_ip,
                "nick": writer_nick
                }

    def get_post_title(self):
        try:
            title = self.post_tr.find(class_="gall_tit ub-word").find("a")
            return {"url": urljoin(GALL_URL, title["href"]),
                    "text": title.text}
        except ValueError:
            logging.warning("오류: 제목을 찾지 못하였습니다")
            logging.warning(f"tr 내용: {self.post_tr}")


class MaxQueue:
    def __init__(self, max_size):
        self._queue = []
        self.max_size = max_size

    def put(self, item):
        self._queue.insert(0, item)
        if len(self._queue) > self.max_size:
            self._queue.pop()

    def __contains__(self, item):
        if item in self._queue:
            return True
        else:
            return False


class PostsCollector:
    def __init__(self, gall_url, posts_queue, shutdown_timer, IP_Distinguisher):
        self.gall_url = gall_url
        self.posts_queue = posts_queue
        self.posts_log = MaxQueue(POST_WATCH_RANGE)
        self.shutdown_timer = shutdown_timer
        self.ip_distinguisher = IP_Distinguisher
        self._is_finished = False
        self.rq_headers = {
            "Referer": "https://www.google.com/",
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:89.0) Gecko/20100101 Firefox/89.0"
        }
        logging.info("PostsCollector 초기화")

    @property
    def is_finished(self):
        return self._is_finished

    def collect_posts(self, thread_id=""):
        logging.info(f"POST THREAD: 시작")
        while self.shutdown_timer.is_before_shutdown_time():
            time.sleep(2)
            logging.info("POST THREAD: 글목록 새로고침")
            html = requests.get(self.gall_url, headers=self.rq_headers).text
            soup = BeautifulSoup(html, "html.parser")
            trs = soup.find_all("tr", class_="ub-content us-post")

            for i in range(POST_WATCH_RANGE - 1, -1, -1):
                tr = trs[i]
                post_data_getter = PostDataGetter(tr)
                post_num = post_data_getter.get_post_num()
                post_type_str = post_data_getter.get_post_type_str()
                writer = post_data_getter.get_post_writer()
                writer_ip = writer["ip"]

                if not writer_ip:
                    writer["type"] = "고닉"

                else:
                    is_korean_ip = self.ip_distinguisher.is_korean_ip(writer_ip)

                    if is_korean_ip:
                        writer["type"] = "국내유동"
                    elif not is_korean_ip:
                        writer["type"] = "해외유동"
                    else:
                        logging.warning("에러: 유저특정 불가")
                        logging.warning(f"아이피: {writer_ip}")
                        raise ValueError

                img_em_classes = ["icon_recomimg", "icon_btimebest", "icon_rtimebest", "icon_pic", "icon_recmovie",
                                  "icon_movie"]

                if post_num not in self.posts_log and post_type_str in img_em_classes \
                        and writer["type"] in WRITER_TYPES:
                    post_to_read = {
                        "title": post_data_getter.get_post_title(),  # text, url
                        "writer": writer  # ip, nick, type
                    }
                    logging.info(f"POST THREAD: PUT: {post_to_read['title']['text']}")
                    self.posts_queue.put(post_to_read)
                    self.posts_log.put(post_num)
                else:
                    logging.debug("읽을 글이 아닌 걸로 판명, 다음글로")

        self._is_finished = True
        logging.info(f"POST THREAD: 끝")


class SrcsCollector:
    def __init__(self, posts_queue, srcs_queue, posts_collector):
        self._is_finished = False
        self.posts_queue = posts_queue
        self.srcs_queue = srcs_queue
        self.posts_collector = posts_collector
        self.post_rq_headers = {
            "Referer": GALL_URL,
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:89.0) Gecko/20100101 Firefox/89.0"
        }
        logging.info("SrcsCollector 초기화")

    @property
    def is_finished(self):
        return self._is_finished

    def collect_srcs(self, thread_id=""):
        logging.info(f"SRCS THREAD {thread_id}: 시작")
        while not self.posts_collector.is_finished or not self.posts_queue.empty():
            if not self.posts_queue.empty():
                post = self.posts_queue.get()
                logging.info(f"SRCS THREAD {thread_id}: GET: {post['title']['text']}")
                html = requests.get(post["title"]["url"], headers=self.post_rq_headers).text
                post_soup = BeautifulSoup(html, "html.parser")
                post_written_time = post_soup.find(class_="gall_date").string
                src_rq_headers = {
                    "Referer": post["title"]["url"],
                    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:89.0) Gecko/20100101 Firefox/89.0"
                }
                post["written_time"] = post_written_time
                post["src_rq_headers"] = src_rq_headers

                wr_div = post_soup.find("div", class_="write_div")

                pics = wr_div.find_all("img")

                vids = wr_div.find_all("video")
                mp4s = []
                gifs = []
                for vid in vids:
                    if vid.has_attr("data-src"):
                        gifs.append(vid)
                    elif vid.has_attr("class"):
                        if 'dc_mv' in vid["class"]:
                            mp4s.append(vid)
                        else:
                            logging.debug("모르는 동영상")

                logging.info(f"SRCS THREAD {thread_id}: 사진: {len(pics)}, 움짤: {len(gifs)}, 동영상: {len(mp4s)} 발견")

                media_srcs = []
                for img in pics:
                    src = img["src"]
                    media_srcs.append(src)
                for img in gifs:
                    src = img["data-src"]
                    media_srcs.append(src)
                for img in mp4s:
                    src = img.source["src"]
                    media_srcs.append(src)

                post["media_srcs"] = media_srcs

                logging.info(f"SRCS THREAD {thread_id}: 미디어 소스 {len(media_srcs)}개 추출")
                self.srcs_queue.put(post)
                logging.info(f"SRCS THREAD {thread_id}: 남은 글 수: {self.posts_queue.qsize()}")
            else:
                time.sleep(3)
            logging.debug(f"SRCS THREAD {thread_id}: while 끝")

        self._is_finished = True
        logging.info(f"SRCS THREAD {thread_id}: 끝")


def replace_forbbiden_char(strr):
    forbbden_chars = ["<", ">", ":", "\"", "/", "\\", "|", "?", "*"]
    ret = strr
    for fc in forbbden_chars:
        ret = ret.replace(fc, "X")

    return ret


def find_parameter(str_params, param_name):
    str_params = str_params.split(";")

    for str_param in str_params:
        if param_name in str_param:
            value = str_param.split("=")[1]
            return value
    return ""


def find_extension(filename):
    if "." in filename:
        extension = filename.split(".")[-1]
        return extension
    else:
        return ""


class Downloader:
    def __init__(self, srcs_queue, srcs_collector):
        self.srcs_queue = srcs_queue
        self.srcs_collector = srcs_collector
        self.download_cnt = 0
        logging.info("Downloader 초기화")

    def download(self, thread_id=""):
        logging.info(f"DOWN THREAD: 시작")
        while not self.srcs_collector.is_finished or not self.srcs_queue.empty():
            logging.debug("download while 시작")

            if not self.srcs_queue.empty():
                post = self.srcs_queue.get()
                media_srcs = post["media_srcs"]
                rq_headers = post["src_rq_headers"]

                full_title = post["title"]["text"]
                sl_title = full_title[0:5] if len(full_title) > 5 else full_title
                rp_title = replace_forbbiden_char(sl_title)
                written_time = post["written_time"].replace(":", ".")


                user_info = ""
                if post["writer"]["type"] == "고닉":
                    user_info = post["writer"]["nick"]
                else:
                    user_info = f"{post['writer']['nick']}({post['writer']['ip']})[{post['writer']['type']}]"
                user_info = replace_forbbiden_char(user_info)
                for i, media_src in enumerate(media_srcs):
                    media_response = requests.get(media_src, headers=rq_headers)

                    extension = ""
                    if "content-disposition" in media_response.headers:
                        disposition_prams = media_response.headers['content-disposition']
                        original_filename = find_parameter(disposition_prams, "filename")
                        extension = find_extension(original_filename)
                    elif "content-type" in media_response.headers:
                        if "mp4" in media_response.headers["content-type"]:
                            extension = "mp4"
                    else:
                        logging.warning("알 수 없는 확장자")
                        logging.warning(f"url: {post['title']['url']}")
                        logging.warning(f"src: {media_src}")
                        raise ValueError

                    save_file_name = f"{rp_title}_{user_info}_{written_time}_{i:03d}{'.' + extension if extension else ''}"
                    save_dir = ""
                    if SAVE_DIR:
                        if not os.path.exists(SAVE_DIR):
                            logging.warning("경로가 존재하지 않습니다.")
                            raise ValueError
                        save_dir = SAVE_DIR
                    else:
                        if os.path.exists(os.path.join(dirname, "images")):
                            pass
                        else:
                            os.mkdir(os.path.join(dirname, "images"))
                        save_dir = os.path.join(dirname, "images")
                    save_path = f"{os.path.join(save_dir, save_file_name)}"
                    try:
                        with open(save_path, "wb") as f:
                            f.write(media_response.content)
                        logging.info(f"DOWN THREAD: {save_file_name} 다운로드")
                        self.download_cnt += 1
                    except OSError:
                        logging.warning("잘못된 파일명")
                    time.sleep(1)

                logging.info(f"DOWN THREAD: 누적 다운로드 수: {self.download_cnt}, 남은 다운로드 수: {self.srcs_queue.qsize()}")
            else:
                time.sleep(5)
            logging.debug("download while 끝")
        logging.info(f"DOWN THREAD: 끝")


class ThreadManager:
    def __init__(self):
        self.post_queue = queue.Queue()
        self.src_queue = queue.Queue()

        self.posts_collector = PostsCollector(GALL_URL, self.post_queue,
                                              ShutdownTimer(SHUTDOWN_TIME, t_delta=SHUTDOWN_TIME_DELTA),
                                              IP_Distinguisher(KOREAN_IPS_XLSX))
        self.srcs_collector = SrcsCollector(self.post_queue, self.src_queue, self.posts_collector)
        self.downloader = Downloader(self.src_queue, self.srcs_collector)

    def start(self):
        with concurrent.futures.ThreadPoolExecutor(max_workers=2 + COLLECTOR_THREADS_NUM) as executor:
            executor.submit(self.posts_collector.collect_posts)
            for i in range(COLLECTOR_THREADS_NUM):
                executor.submit(self.srcs_collector.collect_srcs, i)
            executor.submit(self.downloader.download)


if __name__ == "__main__":
    tm = ThreadManager()
    tm.start()
